"""Generate LoveLab_Order_Template_Pack4.xlsx matching the style of Pack 1-3."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(
    os.path.dirname(__file__),
    "../public/LoveLab Excel Packs/LoveLab_Order_Template_Pack4.xlsx"
)

# ── Brand colours ────────────────────────────────────────────────────────────
PLUM      = "5D3A5E"
PLUM_DARK = "4A2545"
WHITE     = "FFFFFF"
LIGHT_ROW = "FFFFF9FF"  # very faint lavender
ALT_ROW   = "F5EFF5"
TEXT_GRAY = "4F4F4F"
MUTED     = "8A6A7D"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=TEXT_GRAY, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="middle", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def thin_border(color="D0C0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_top_border(top_color=PLUM, rest_color="D0C0D0"):
    return Border(
        top=Side(style="medium", color=top_color),
        left=Side(style="thin", color=rest_color),
        right=Side(style="thin", color=rest_color),
        bottom=Side(style="thin", color=rest_color),
    )

# ── Pack 4 data ──────────────────────────────────────────────────────────────
PACK4_ROWS = [
    ("SHAPY SHINE FANCY", "0.10 ct", "Heart",    "Yellow", "Bezel", "M",   "Grey",        1, 55,  "IGI"),
    ("SHAPY SHINE FANCY", "0.10 ct", "Pear",     "Yellow", "Bezel", "M",   "Red",         1, 55,  "IGI"),
    ("SHAPY SHINE FANCY", "0.10 ct", "Marquise", "White",  "Bezel", "M",   "Navy Blue",   1, 55,  "IGI"),
    ("SHAPY SHINE FANCY", "0.10 ct", "Emerald",  "Yellow", "Bezel", "M",   "Black",       1, 55,  "IGI"),
    ("SHAPY SHINE FANCY", "0.10 ct", "Oval",     "White",  "Bezel", "M",   "Bordeaux",    1, 55,  "IGI"),
    ("SHAPY SHINE FANCY", "0.30 ct", "Oval",     "White",  "Prong", "M",   "Gold",        1, 100, "IGI"),
    ("SHAPY SHINE FANCY", "0.30 ct", "Emerald",  "Yellow", "Prong", "M",   "Lilac",       1, 100, "IGI"),
    ("SHAPY SHINE FANCY", "0.30 ct", "Pear",     "White",  "Prong", "M",   "Light Pink",  1, 100, "IGI"),
    ("MULTI FOUR",        "0.20 ct", "",         "White",  "",      "M",   "Gold",        1, 85,  "IGI"),
    ("MULTI FOUR",        "0.20 ct", "",         "Yellow", "",      "M",   "Black",       1, 85,  "IGI"),
    ("MULTI FOUR",        "0.20 ct", "",         "Yellow", "",      "M",   "Bordeaux",    1, 85,  "IGI"),
    ("MULTI THREE",       "0.15 ct", "",         "",       "LO",    "",    "Gold",        1, 65,  "IGI"),
    ("MULTI THREE",       "0.15 ct", "",         "WWW",    "F",     "",    "Black",       1, 65,  "IGI"),
    ("MULTI THREE",       "0.15 ct", "",         "YYY",    "F",     "",    "Bordeaux",    1, 65,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "Yellow", "",      "S/M", "Red",         1, 30,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "Yellow", "",      "S/M", "Bordeaux",    1, 30,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "White",  "",      "S/M", "Gold",        1, 30,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "White",  "",      "S/M", "Black",       1, 30,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "Yellow", "",      "S/M", "Silver Grey", 1, 30,  "IGI"),
    ("CUBIX",             "0.05 ct", "",         "White",  "",      "S/M", "Navy Blue",   1, 30,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "Yellow", "",      "S/M", "Red",         1, 40,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "Yellow", "",      "S/M", "Bordeaux",    1, 40,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "White",  "",      "S/M", "Gold",        1, 40,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "White",  "",      "S/M", "Black",       1, 40,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "Yellow", "",      "S/M", "Silver Grey", 1, 40,  "IGI"),
    ("CUBIX",             "0.10 ct", "",         "White",  "",      "S/M", "Navy Blue",   1, 40,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "White",  "",      "M",   "Gold",        1, 30,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "Yellow", "",      "M",   "Silver Grey", 1, 30,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "White",  "",      "M",   "Black",       1, 30,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "White",  "",      "M",   "Navy Blue",   1, 30,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "Yellow", "",      "M",   "Red",         1, 30,  "IGI"),
    ("CUTY",              "0.05 ct", "",         "Yellow", "",      "M",   "Bordeaux",    1, 30,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "White",  "",      "M",   "Gold",        1, 40,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "Yellow", "",      "M",   "Silver Grey", 1, 40,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "White",  "",      "M",   "Black",       1, 40,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "White",  "",      "M",   "Navy Blue",   1, 40,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "Yellow", "",      "M",   "Red",         1, 40,  "IGI"),
    ("CUTY",              "0.10 ct", "",         "Yellow", "",      "M",   "Bordeaux",    1, 40,  "IGI"),
]

HEADERS = ["Collection", "Carat", "Shape", "Housing", "Setting", "Size",
           "Cord Color", "Qty", "Unit Price €", "Total €", "Cert", "Reference", "Notes"]
COL_WIDTHS = [24, 10, 14, 12, 12, 8, 16, 7, 14, 13, 10, 22, 26]
NUM_COLS = len(HEADERS)
LAST_COL = get_column_letter(NUM_COLS)

def write():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pack 4 Order"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze below header row (row 8)
    ws.freeze_panes = "A9"

    # ── Row 1: Brand title ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 44
    ws.merge_cells(f"A1:{LAST_COL}1")
    c = ws["A1"]
    c.value = "✦  LoveLab"
    c.font  = Font(name="Calibri", bold=True, size=20, color=WHITE)
    c.fill  = fill(PLUM)
    c.alignment = align("left", "center", indent=2)

    # ── Row 2: Pack subtitle ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A2:{LAST_COL}2")
    c = ws["A2"]
    c.value = "Order Template — Pack 4   ·   SHAPY SHINE FANCY  ·  MULTI FOUR  ·  MULTI THREE  ·  CUBIX  ·  CUTY"
    c.font  = Font(name="Calibri", size=10, color="CFAECF", italic=True)
    c.fill  = fill(PLUM)
    c.alignment = align("left", "center", indent=2)

    # ── Row 3: Plum spacer ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6
    ws.merge_cells(f"A3:{LAST_COL}3")
    ws["A3"].fill = fill(PLUM)

    # ── Row 4: Client info labels ────────────────────────────────────────────
    ws.row_dimensions[4].height = 16
    info_labels = [("A4", "D4", "CLIENT / COMPANY"),
                   ("E4", "H4", "CONTACT NAME"),
                   ("I4", "J4", "DATE"),
                   ("K4", "M4", "PO / REFERENCE")]
    for start, end, label in info_labels:
        ws.merge_cells(f"{start}:{end}")
        c = ws[start]
        c.value = label
        c.font  = Font(name="Calibri", size=8, color=MUTED, bold=True)
        c.fill  = fill("FFF7FF")
        c.alignment = align("left", "center", indent=1)

    # ── Row 5: Client info input cells ──────────────────────────────────────
    ws.row_dimensions[5].height = 24
    info_ranges = [("A5","D5"), ("E5","H5"), ("I5","J5"), ("K5","M5")]
    for start, end in info_ranges:
        ws.merge_cells(f"{start}:{end}")
        c = ws[start]
        c.fill   = fill("FBF8FF")
        c.border = thin_border("D8C8D8")
        c.font   = font(size=11)
        c.alignment = align("left", "center", indent=1)

    # ── Row 6: Spacer ────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 8
    ws.merge_cells(f"A6:{LAST_COL}6")
    ws["A6"].fill = fill("F0E8F0")

    # ── Row 7: Instructions ──────────────────────────────────────────────────
    ws.row_dimensions[7].height = 16
    ws.merge_cells(f"A7:{LAST_COL}7")
    c = ws["A7"]
    c.value = "Fill in Qty and Reference columns. All other fields are pre-filled for Pack 4. Send the completed form to your LoveLab representative."
    c.font  = Font(name="Calibri", size=9, color=MUTED, italic=True)
    c.fill  = fill("F7F0F7")
    c.alignment = align("left", "center", indent=2)

    # ── Row 8: Column headers ────────────────────────────────────────────────
    ws.row_dimensions[8].height = 26
    for col_idx, header in enumerate(HEADERS, 1):
        c = ws.cell(row=8, column=col_idx, value=header)
        c.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill  = fill(PLUM)
        c.alignment = align("center", "center", wrap=True)
        s = Side(style="thin", color="7A4F7C")
        c.border = Border(left=s, right=s, top=s, bottom=s)

    # ── Data rows ────────────────────────────────────────────────────────────
    prev_collection = None
    for i, row_data in enumerate(PACK4_ROWS):
        row_num = i + 9
        is_alt = (i % 2 == 1)
        bg = ALT_ROW if is_alt else "FFFFF9"
        is_new_section = row_data[0] != prev_collection
        prev_collection = row_data[0]

        col_idx_map = row_data  # tuple: collection, carat, shape, housing, setting, size, cord, qty, unit_price, cert
        values = list(row_data[:9]) + [None, row_data[9], "", ""]
        # Total column: formula
        total_formula = f"=H{row_num}*I{row_num}"

        for ci, val in enumerate(values, 1):
            if ci == 10:
                c = ws.cell(row=row_num, column=ci)
                c.value = total_formula
                c.number_format = '€#,##0.00'
            else:
                c = ws.cell(row=row_num, column=ci, value=val)

            c.fill = fill(bg)
            c.font = font(size=10)

            if ci in (8, 9, 10):
                c.alignment = align("center", "center")
                if ci == 9:
                    c.number_format = '€#,##0.00'
            else:
                c.alignment = align("left", "center", indent=1)

            if is_new_section and ci > 0:
                c.border = medium_top_border()
            else:
                c.border = thin_border("E0D0E0")

        ws.row_dimensions[row_num].height = 20

    # ── Totals row ───────────────────────────────────────────────────────────
    total_row = 9 + len(PACK4_ROWS)
    ws.row_dimensions[total_row].height = 28

    ws.merge_cells(f"A{total_row}:G{total_row}")
    c = ws[f"A{total_row}"]
    c.value = "TOTAL ORDER VALUE"
    c.font  = Font(name="Calibri", bold=True, size=12, color=WHITE)
    c.fill  = fill(PLUM_DARK)
    c.alignment = align("right", "center", indent=2)

    # Qty total
    c = ws[f"H{total_row}"]
    c.value = f"=SUM(H9:H{total_row - 1})"
    c.font  = Font(name="Calibri", bold=True, size=12, color=WHITE)
    c.fill  = fill(PLUM_DARK)
    c.alignment = align("center", "center")

    # Unit price cell (blank, colored)
    ws[f"I{total_row}"].fill = fill(PLUM_DARK)

    # Grand total
    c = ws[f"J{total_row}"]
    c.value = f"=SUM(J9:J{total_row - 1})"
    c.number_format = '€#,##0.00'
    c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill  = fill(PLUM_DARK)
    c.alignment = align("center", "center")
    s = Side(style="medium", color=WHITE)
    c.border = Border(left=s, right=s, top=s, bottom=s)

    for col_idx in range(11, NUM_COLS + 1):
        ws.cell(row=total_row, column=col_idx).fill = fill(PLUM_DARK)

    # ── Footer ───────────────────────────────────────────────────────────────
    footer_row = total_row + 2
    ws.merge_cells(f"A{footer_row}:{LAST_COL}{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = "LoveLab  ·  hello@love-lab.com  ·  Generated by LoveLab B2B Platform"
    c.font  = Font(name="Calibri", size=8, color="CCAACC", italic=True)
    c.alignment = align("center", "center")

    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"✅  Written: {OUT}")

write()
